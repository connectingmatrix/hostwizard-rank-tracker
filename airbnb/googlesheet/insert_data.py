import re
import sys
import time
import gspread
import requests
import calendar
import warnings
import pandas as pd
from datetime import date
from bs4 import BeautifulSoup
from difflib import SequenceMatcher
from selenium import webdriver
from gspread_formatting import *
warnings.filterwarnings("ignore")
from gspread.exceptions import APIError
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl.utils import column_index_from_string, get_column_letter
from requests.exceptions import ConnectionError, Timeout, TooManyRedirects
from selenium.common.exceptions import TimeoutException, WebDriverException, InvalidSessionIdException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

def background(f):
    def wrapped(*args, **kwargs):
        return asyncio.get_event_loop().run_in_executor(None, f, *args, **kwargs)

    return wrapped

class DataInsertion:

    scope = ['https://spreadsheets.google.com/feeds',
             'https://www.googleapis.com/auth/spreadsheets',
             'https://www.googleapis.com/auth/drive.file',
             'https://www.googleapis.com/auth/drive']

    RULES_SEARCH_PAGE = {  
    'Name': {'tag': 'div', 'class': 't1jojoys'},
    'Url': {'tag': 'a', 'get': 'href'},
    'Description': {'tag': 'div', 'class': 'fb4nyux'},
    'Rooms and Beds': {'tag': 'div', 'class': 'fb4nyux','order':1},
    #'Price': {'tag': 'div', 'class': '_1jo4hgw', },
    'Total_Price': {'tag': 'div', 'class': '_tt122m'},
    }

    # Reading Credentails from ServiceAccount Keys file
    credentials = ServiceAccountCredentials.from_json_keyfile_name('googlesheet/ranking-automation-33939eb3373d.json', scope)
    client = gspread.authorize(credentials)
    # Build the Drive API client
    drive_service = build('drive', 'v3', credentials=credentials)
    sheets_service = build('sheets', 'v4', credentials=credentials)
    
    driver = None

    def print(*args, **kwargs):
        print(*args, **kwargs)
        sys.stdout.flush()
        
    def __int__(self):
        #self.setup_driver()
        
        self.rank_status= False

    def setup_driver(self):
        
        if self.driver is None:
            options = Options()

            options.add_argument('--disable-extensions')
            options.add_argument('--headless')
            options.add_argument('--no-sandbox')
            options.add_argument('--disable-dev-shm-usage')
            # options.add_argument('--remote-debugging-port=9515')
            options.add_argument('--disable-setuid-sandbox')
            options.add_argument("--ignore-certificate-errors")
            options.add_argument("--ignore-ssl-errors")
            #options.AcceptInsecureCertificates = True
            options.accept_insecure_certs= True
            #options.add_argument('--incognito')

            driver = webdriver.Chrome(ChromeDriverManager().install(), options=options)
            # driver = webdriver.Chrome(options=options)
            self.driver = driver
        else:
            self.print('Driver is present')

    def get_spreadsheet_id(self, spreadsheet_name):
        success = False
        while not success:
            # Define the search query to find the workbook
            query = f"name='{spreadsheet_name}' and mimeType='application/vnd.google-apps.spreadsheet' and trashed=false"

            # Call the Drive API to search for the workbook
            try:
                results = self.drive_service.files().list(q=query).execute()
                items = results.get('files', [])

                if len(items) == 0:
                    self.print("Workbook not found")
                    workbook_id = 0
                else:
                    # The workbook is found, continue with the script
                    workbook_id = items[0]['id']
                
                success = True
                return workbook_id
            
            except APIError as error:
                
                if error.response.status_code == 429:
                    self.print("Quota exceeded. Retrying in 40 seconds...")
                    time.sleep(40)  # wait for 40 seconds
                else:
                    self.print("APIError occurred: %s" % error)
                    self.print("Retrying in 10 seconds...")
                    time.sleep(15)  # wait for 10 seconds

            except (ConnectionError, Timeout, TooManyRedirects) as error:
                self.print("Connection error occurred: %s" % error)
                self.print("Retrying in 10 seconds...")
                time.sleep(10)  # wait for 10 seconds

            except Exception as error:
                self.print("An error occurred: %s" % error)
                self.print("Retrying in 10 seconds...")
                time.sleep(10)  # wait for 10 seconds

    

    def get_or_create_sheet_id(self, spreadsheet_name, sheet_name):
        success = False
        while not success:
            try:
                spreadsheet_id = self.get_spreadsheet_id(spreadsheet_name)
                sheet_metadata = self.sheets_service.spreadsheets().get(
                    spreadsheetId=spreadsheet_id).execute()
                sheets = sheet_metadata.get('sheets', [])

                for sheet in sheets:
                    if sheet['properties']['title'] == sheet_name:
                        # The sheet already exists, get its ID
                        sheet_id = sheet['properties']['sheetId']
        
                        return sheet_id

                # The sheet does not exist, create it and get its ID
                self.print('Sheet is not present, now creating new Sheet for this hotel..')
                requests = [{
                    "addSheet": {
                        "properties": {
                            "title": sheet_name
                        }
                    }
                }]
                body = {
                    "requests": requests
                }
                response = self.sheets_service.spreadsheets().batchUpdate(
                    spreadsheetId=spreadsheet_id, body=body).execute()
                sheet_id = response['replies'][0]['addSheet']['properties']['sheetId']
                self.print( f"A new worksheet named '{sheet_name}' has been added to the Google Spreadsheet with ID '{sheet_id}'")
                success= True
                return sheet_id

            except APIError as error:
                
                if error.response.status_code == 429:
                    self.print("Quota exceeded. Retrying in 40 seconds...")
                    time.sleep(40)  # wait for 40 seconds
                else:
                    self.print("APIError occurred: %s" % error)
                    self.print("Retrying in 10 seconds...")
                    time.sleep(15)  # wait for 10 seconds

            except (ConnectionError, Timeout, TooManyRedirects) as error:
                self.print("Connection error occurred: %s" % error)
                self.print("Retrying in 10 seconds...")
                time.sleep(10)  # wait for 10 seconds

            except Exception as error:
                self.print("An error occurred: %s" % error)
                self.print("Retrying in 10 seconds...")
                time.sleep(10)  # wait for 10 seconds
            

    def insert_cells_at_start(self, spreadsheet_name, sheet_name, num_columns):
        success = False
        while not success:
            try:
                # Get the spreadsheet ID
                spreadsheet_id = self.get_spreadsheet_id(spreadsheet_name)

                # Get the sheet ID
                sheet_id = self.get_or_create_sheet_id(spreadsheet_name,sheet_name)

                # Define the cells to insert
                cells = []
                for i in range(num_columns):
                    cells.append({})

                # Insert the cells at the beginning of the worksheet
                self.sheets_service.spreadsheets().batchUpdate(
                    spreadsheetId=spreadsheet_id,
                    body={
                        'requests': [
                            {
                                'insertDimension': {
                                    'range': {
                                        'sheetId': sheet_id,
                                        'dimension': 'COLUMNS',
                                        'startIndex': 0,
                                        'endIndex': num_columns
                                    },
                                    'inheritFromBefore': False
                                }
                            },
                            {
                                'updateCells': {
                                    'rows': [
                                        {
                                            'values': cells
                                        }
                                    ],
                                    'fields': 'userEnteredValue',
                                    'start': {
                                        'sheetId': sheet_id,
                                        'rowIndex': 0,
                                        'columnIndex': 0
                                    }
                                }
                            }
                        ]
                    }
                ).execute()
                success = True

            except APIError as error:
                
                if error.response.status_code == 429:
                    self.print("Quota exceeded. Retrying in 40 seconds...")
                    time.sleep(60)  # wait for 40 seconds
                else:
                    self.print("APIError occurred: %s" % error)
                    self.print("Retrying in 10 seconds...")
                    time.sleep(10)  # wait for 10 seconds

            except (ConnectionError, Timeout, TooManyRedirects) as error:
                self.print("Connection error occurred: %s" % error)
                self.print("Retrying in 10 seconds...")
                time.sleep(10)  # wait for 10 seconds

            except Exception as error:
                self.print("An error occurred: %s" % error)
                self.print("Retrying in 10 seconds...")
                time.sleep(10)  # wait for 10 seconds


    def update_cell_size(self, spreadsheet_name, sheet_name, cell_range, row_height, col_width):
        # Get the spreadsheet ID
        spreadsheet_id = self.get_spreadsheet_id(spreadsheet_name)

        # Get the sheet ID
        sheet_id = self.get_or_create_sheet_id(sheet_name)

        # create a request body that specifies the new height and width of the cell(s)
        requests = [{
            'updateDimensionProperties': {
                'range': {
                    'sheetId': sheet_id,
                    'dimension': 'ROWS',
                    'startIndex': int(cell_range.split(':')[0][1:]) - 1,
                    'endIndex': int(cell_range.split(':')[1][1:])
                },
                'properties': {
                    'pixelSize': row_height
                },
                'fields': 'pixelSize'
            }
        },
            {
                'updateDimensionProperties': {
                    'range': {
                        'sheetId': sheet_id,
                        'dimension': 'COLUMNS',
                        'startIndex': ord(cell_range.split(':')[0][0]) - 65,
                        'endIndex': ord(cell_range.split(':')[1][0]) - 64
                    },
                    'properties': {
                        'pixelSize': col_width
                    },
                    'fields': 'pixelSize'
                }
            }]

        # execute the request to update the cell(s) with the new height and width
        try:
            self.sheets_service.spreadsheets().batchUpdate(spreadsheetId=spreadsheet_id,
                                                           body={'requests': requests}).execute()
            self.print(
                f"Cell(s) {cell_range} in sheet '{sheet_name}' of spreadsheet '{spreadsheet_id}' have been updated with height {row_height} and width {col_width}.")
        
        except APIError as error:
                
            if error.response.status_code == 429:
                self.print("Quota exceeded. Retrying in 40 seconds...")
                time.sleep(40)  # wait for 40 seconds
            else:
                self.print("APIError occurred: %s" % error)
                self.print("Retrying in 10 seconds...")
                time.sleep(10)  # wait for 10 seconds

        except (ConnectionError, Timeout, TooManyRedirects) as error:
            self.print("Connection error occurred: %s" % error)
            self.print("Retrying in 10 seconds...")
            time.sleep(10)  # wait for 10 seconds

        except Exception as error:
            self.print("An error occurred: %s" % error)
            self.print("Retrying in 10 seconds...")
            time.sleep(10)  # wait for 10 seconds

    def combine_and_update_col(self, spreadsheet_name, sheet_name, column_range, value):
        success = False
        while not success:
            try:
                # Open the desired spreadsheet by its name
                sheet = self.client.open(spreadsheet_name).worksheet(sheet_name)

                # Get the cell range object for the specified range
                cell_range = sheet.range(column_range)

                # Set the value in the first cell of the range
                cell_range[0].value = value

                # Merge the remaining cells in the range into a single cell
                sheet.merge_cells(column_range)

                # Update the first cell in the range with the new value
                sheet.update_cells(cell_range)
                success = True

            except APIError as error:
                
                if error.response.status_code == 429:
                    self.print("Quota exceeded. Retrying in 40 seconds...")
                    time.sleep(40)  # wait for 40 seconds
                else:
                    self.print("APIError occurred: %s" % error)
                    self.print("Retrying in 10 seconds...")
                    time.sleep(10)  # wait for 10 seconds

            except (ConnectionError, Timeout, TooManyRedirects) as error:
                self.print("Connection error occurred: %s" % error)
                self.print("Retrying in 10 seconds...")
                time.sleep(10)  # wait for 10 seconds

            except Exception as error:
                self.print("An error occurred: %s" % error)
                self.print("Retrying in 10 seconds...")
                time.sleep(10)  # wait for 10 seconds


    def update_spreadsheet_cell(self, spreadsheet_name, sheet_name, cell, value):
        self.print(cell)
        success = False
        while not success:
            # Open the desired spreadsheet by its name
            try:
                sheet = self.client.open(spreadsheet_name).worksheet(sheet_name)
                # Define the cell where you want to add data and the value to be added
                col, row = cell[0], int(cell[1:])
                #cell = sheet.cell(row, ord(col) - 64)

                # Update the cell with the new value
                sheet.update_cell(row, ord(col) - 64, value)
                success = True  # set success flag to true

            except APIError as error:
                
                if error.response.status_code == 429:
                    self.print("Quota exceeded. Retrying in 40 seconds...")
                    time.sleep(40)  # wait for 40 seconds
                    continue
                else:
                    self.print("APIError occurred: %s" % error)
                    self.print("Retrying in 10 seconds...")
                    time.sleep(40)  # wait for 10 seconds
                    continue

            except (ConnectionError, Timeout, TooManyRedirects) as error:
                self.print("Connection error occurred: %s" % error)
                self.print("Retrying in 10 seconds...")
                time.sleep(20)  # wait for 10 seconds
                continue

            except Exception as error:
                
                self.print("An error occurred: %s" % error)
                self.print("Retrying in 10 seconds...")
                time.sleep(10)  # wait for 10 seconds
                continue
            


    def get_color_rgb(self, color_name):
        if color_name == 'white':
            return 1, 1, 1
        elif color_name == 'green':
            return 0, 1, 0
        elif color_name == 'blue':
            return 0.41, 0.619, 0.90
        else:
            raise ValueError('Invalid color name')

    def update_cell_color(self, spreadsheet_name, sheet_name, cell_coordinates, color_name):
        # Get the ID of the "Ranking" spreadsheet and the ID of the "Sheet10" sheet
        spreadsheet_id = self.get_spreadsheet_id(spreadsheet_name)
        sheet_id = self.get_or_create_sheet_id(spreadsheet_name, sheet_name)

        # Convert cell coordinates to row and column indices 
        row, col = gspread.utils.a1_to_rowcol(cell_coordinates)

        # Convert color name to RGB values
        color = self.get_color_rgb(color_name)

        # Define the API request to update the cell color
        batch_update_spreadsheet_request_body = {
            "requests": [
                {
                    "repeatCell": {
                        "range": {
                            "sheetId": sheet_id,
                            "startRowIndex": row - 1,
                            "endRowIndex": row,
                            "startColumnIndex": col - 1,
                            "endColumnIndex": col
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColor": {
                                    "red": color[0],
                                    "green": color[1],
                                    "blue": color[2]
                                }
                            }
                        },
                        "fields": "userEnteredFormat.backgroundColor"
                    }
                }
            ]
        }
        # Send the API request to update the cell color
        success = False
        while not success:
            try:
                request = self.sheets_service.spreadsheets().batchUpdate(spreadsheetId=spreadsheet_id,
                                    body=batch_update_spreadsheet_request_body)
                request.execute()
                success = True

            except APIError as error:
                
                if error.response.status_code == 429:
                    self.print("Quota exceeded. Retrying in 40 seconds...")
                    time.sleep(40)  # wait for 40 seconds
                else:
                    self.print("APIError occurred: %s" % error)
                    self.print("Retrying in 10 seconds...")
                    time.sleep(10)  # wait for 10 seconds

            except (ConnectionError, Timeout, TooManyRedirects) as error:
                self.print("Connection error occurred: %s" % error)
                self.print("Retrying in 10 seconds...")
                time.sleep(10)  # wait for 10 seconds

            except Exception as error:
                self.print("An error occurred: %s" % error)
                self.print("Retrying in 10 seconds...")
                time.sleep(10)  # wait for 10 seconds

    def add_checkbox(self, spreadsheet_name,  sheet_name, cell):

        spreadsheet_id = self.get_spreadsheet_id(spreadsheet_name)

        # Open the specified spreadsheet and worksheet
        spreadsheet = self.client.open_by_key(spreadsheet_id)
        sheet = spreadsheet.worksheet(sheet_name)
        sheet_id = sheet._properties['sheetId']

        # Parse the cell address to get the row and column numbers
        row, column = gspread.utils.a1_to_rowcol(cell)

        # Build the requests to add a checkbox to the specified cell
        requests = {
            "requests": [
                {
                    "repeatCell": {
                        "cell": {"dataValidation": {"condition": {"type": "BOOLEAN"}}},
                        "range": {"sheetId": sheet_id, "startRowIndex": row - 1, "endRowIndex": row,
                                  "startColumnIndex": column - 1, "endColumnIndex": column},
                        "fields": "dataValidation"
                    }
                },
                {
                    "updateCells": {
                        "rows": [{"values": [{"userEnteredValue": {"boolValue": True}}]}],
                        "start": {"rowIndex": row - 1, "columnIndex": column - 1, "sheetId": sheet_id},
                        "fields": "userEnteredValue"
                    }
                }
            ]
        }
        success = False
        while not success:
            # Open the desired spreadsheet by its name
            try:
                # Execute the requests to add a checkbox to the specified cell
                spreadsheet.batch_update(requests)
                success = True

            except APIError as error:
                
                if error.response.status_code == 429:
                    self.print("Quota exceeded. Retrying in 40 seconds...")
                    time.sleep(40)  # wait for 40 seconds
                else:
                    self.print("APIError occurred: %s" % error)
                    self.print("Retrying in 10 seconds...")
                    time.sleep(10)  # wait for 10 seconds

            except (ConnectionError, Timeout, TooManyRedirects) as error:
                self.print("Connection error occurred: %s" % error)
                self.print("Retrying in 10 seconds...")
                time.sleep(10)  # wait for 10 seconds

            except Exception as error:
                self.print("An error occurred: %s" % error)
                self.print("Retrying in 10 seconds...")
                time.sleep(10)  # wait for 10 seconds
            

    def update_spreadsheet_cells_ranking(self, spreadsheet_name, sheet_name):
        
        self.sheets_service = build('sheets', 'v4', credentials=self.credentials)
        range_and_values = [
            {
                'range': f'{sheet_name}!A1',
                'values': [['DATE OF CHECK IN']]
            },
            {
                'range': f'{sheet_name}!A3',
                'values': [['Date Use']]
            },
            {
                'range': f'{sheet_name}!B3',
                'values': [['Checked Guesty?']]
            },
            {
                'range': f'{sheet_name}!C3',
                'values': [['8']]
            },
            {
                'range': f'{sheet_name}!D3',
                'values': [['9']]
            },
            {
                'range': f'{sheet_name}!E3',
                'values': [['10']]
            },
            {
                'range': f'{sheet_name}!F3',
                'values': [['11']]
            },
            {
                'range': f'{sheet_name}!G3',
                'values': [['12']]
            },
            {
                'range': f'{sheet_name}!H3',
                'values': [['13']]
            },
            {
                'range': f'{sheet_name}!I3',
                'values': [['14']]
            },
            {
                'range': f'{sheet_name}!J3',
                'values': [['15']]
            },
            {
                'range': f'{sheet_name}!K3',
                'values': [['16']]
            }
        ]

        body = {
            'value_input_option': 'USER_ENTERED',
            'data': range_and_values,
        }
        success = False
        while not success:
            try:
                response = self.sheets_service.spreadsheets().values().batchUpdate(
                    spreadsheetId=self.get_spreadsheet_id(spreadsheet_name),
                    body=body,
                ).execute()
                success = True

            except APIError as error:
                
                if error.response.status_code == 429:
                    self.print("Quota exceeded. Retrying in 40 seconds...")
                    time.sleep(40)  # wait for 40 seconds
                else:
                    self.print("APIError occurred: %s" % error)
                    self.print("Retrying in 10 seconds...")
                    time.sleep(10)  # wait for 10 seconds

            except (ConnectionError, Timeout, TooManyRedirects) as error:
                self.print("Connection error occurred: %s" % error)
                self.print("Retrying in 10 seconds...")
                time.sleep(10)  # wait for 10 seconds

            except Exception as error:
                self.print("An error occurred: %s" % error)
                self.print("Retrying in 10 seconds...")
                time.sleep(10)  # wait for 10 seconds

        return response
    

    def Property_Price(self, df, rank, hotel_name,description):
        
        price = df.loc[df['Rank'] == rank, 'Total_Price'].values[0]
        return price
        
    
    def total_hotels(sef, url):

        response = requests.get(url)
        html_content = response.content
        soup = BeautifulSoup(html_content, 'html.parser')
        class_name = 'ty0s4mr'
        element = soup.find(class_=class_name)
        if element is not None:
            matches = re.findall(r'\d+', element.text)
            result = int(''.join(matches))
            return result
#         self.print(result)
        else:
            self.print(f"No element found with class name '{class_name}'")

    def get_all_listings_Hotel_10(self,location, checkin, checkout, guests):

        max_retries = 5
        retries = 0

        while retries < max_retries:
            try:
                self.driver = None
                self.setup_driver()

                # Build the URL
                url = f"https://www.airbnb.com/s/{location}/homes?tab_id=home_tab&refinement_paths%5B%5D=%2Fhomes&flexible_trip_lengths%5B%5D=one_week&price_filter_input_type=0&price_filter_num_nights=3&channel=EXPLORE&date_picker_type=calendar&checkin={checkin}&checkout={checkout}&adults={guests}&source=structured_search_input_header&search_type=search_query"

                wait = WebDriverWait(self.driver, 500)
                locator_strategy = (By.CSS_SELECTOR, 'div.cy5jw6o.dir.dir-ltr')
                self.driver.get(url)
                wait.until(EC.visibility_of_element_located(locator_strategy))

                html_content = self.driver.page_source
                soup = BeautifulSoup(html_content, 'html.parser')

                element = soup.find(class_='tyi4kqb')
                matches = re.findall(r'\d+', element.text)
                hotels = int(''.join(matches))

                total_pages = min(((hotels + 18) - 1) // 18, 3)

                all_listings = []
                if hotels > 0:
                    # Extract all listings from the page
                    for i in range(total_pages):
                        offset = 18 * i
                        new_url = url + f'&items_offset={offset}&section_offset=3'
                        self.driver.get(new_url)
                        wait.until(EC.visibility_of_element_located(locator_strategy))

                        html = self.driver.page_source
                        soup = BeautifulSoup(html, 'html.parser')
                        new_listings = soup.find_all('div', {'class': 'cy5jw6o'})
                        all_listings.extend(new_listings)

                

            except TimeoutException:
                self.print("Timeout exception occurred! Retrying in 10 seconds...")
                self.driver.quit()
                time.sleep(10)
                retries += 1
                continue
            
            except InvalidSessionIdException:
                self.print("Invalid session ID exception occurred! Retrying in 10 seconds...")
                # Re-initialize the driver and wait before trying again
                self.driver.quit()
                time.sleep(10)
                retries += 1
                continue
            
            except Exception as ex:
                self.print(f"Exception occurred: {ex}")
                self.driver.quit()
                time.sleep(5)
                retries += 1
            else:
                break

            finally:

                if self.driver is not None:
                    self.driver.quit()
                    self.driver = None

        return all_listings



    def get_all_listings(self, location, checkin, checkout, guests, hotel_name, description):
        self.rank_status= False
        df_final = pd.DataFrame()
        while True:
            try:
                self.driver = None
                self.setup_driver()
                
                # Build the URL
                url = f"https://www.airbnb.com/s/{location}/homes?tab_id=home_tab&refinement_paths%5B%5D=%2Fhomes&flexible_trip_lengths%5B%5D=one_week&price_filter_input_type=0&price_filter_num_nights=3&channel=EXPLORE&date_picker_type=calendar&checkin={checkin}&checkout={checkout}&adults={guests}&source=structured_search_input_header&search_type=search_query"

                self.driver.get(url)
                wait = WebDriverWait(self.driver, 500)

                locator_strategy = (By.CSS_SELECTOR, 'div.cy5jw6o.dir.dir-ltr')

                wait.until(EC.visibility_of_element_located(locator_strategy))

                html_content = self.driver.page_source
                soup = BeautifulSoup(html_content, 'html.parser')

                element = soup.find(class_='tyi4kqb')
                matches = re.findall(r'\d+', element.text)
                hotels = int(''.join(matches))

                self.print('Hotels', hotels)
                total_pages = min(((hotels + 18) - 1) // 18, 5)
                self.print('Total number of pages: ', total_pages)

                if hotels > 0:
                    df_final = pd.DataFrame()
                    all_listings = []
                    for i in range(total_pages):
                        offset = 18 * i
                        new_url = url + f'&items_offset={offset}&section_offset=3'
                        self.driver.get(new_url)
                        wait.until(EC.visibility_of_element_located(locator_strategy))
                        self.print('Getting Hotels of Page: ', i+1)
                        html = self.driver.page_source
                        soup = BeautifulSoup(html, 'html.parser')
                        new_listings = soup.find_all('div', {'class': 'cy5jw6o'})
                        all_listings.extend(new_listings)

                        df = self.dataframe_buliding(all_listings, self.RULES_SEARCH_PAGE)
                        df_final = pd.concat([df_final, df])
                        df_final = df_final.drop_duplicates()
                        rank = self.Rank_find(df_final, hotel_name, description)
                        if rank > 0:
                            self.rank_status= True
                            break
                
            except TimeoutException:
                self.print("Timeout exception occurred! Retrying in 10 seconds...")
                # Refresh the page and wait for 10 seconds before trying again
                self.driver.quit()
                time.sleep(10)
                continue

            except InvalidSessionIdException:
                self.print("Invalid session ID exception occurred! Retrying in 10 seconds...")
                # Re-initialize the driver and wait before trying again
                self.driver.quit()
                time.sleep(15)
                continue

            except Exception as ex:
                self.print(f"Exception occurred: {ex}")
                self.driver.quit()
                time.sleep(5)
                continue
            
            else:
                break
        
            finally:
                
                if self.driver is not None:
                    self.driver.quit()
                    self.driver = None

        return df_final


    def extract_element(self, listing_html, params):
        # 1. Find the right tag
        if 'class' in params:
            elements_found = listing_html.find_all(params['tag'], params['class'])
        else:
            elements_found = listing_html.find_all(params['tag'])

        # 2. Extract the right element
        tag_order = params.get('order', 0)
        element = elements_found[tag_order]

        # 3. Get text
        if 'get' in params:
            output = element.get(params['get'])
        else:
            output = element.get_text()

        return output

    def dataframe_buliding(self, all_listings, RULES_SEARCH_PAGE):
        # Let's put all the data in to dataframe
        rank = []
        name = []
        description = []
        rooms_and_beds = []
        #price = []
        total_price = []
        link = []
        for i in range(len(all_listings)):
            rank.append(i + 1)
            name.append(self.extract_element(all_listings[i], RULES_SEARCH_PAGE['Name'])),
            link.append(self.extract_element(all_listings[i], RULES_SEARCH_PAGE['Url'])),
            description.append(self.extract_element(all_listings[i], RULES_SEARCH_PAGE['Description'])),
            rooms_and_beds.append(self.extract_element(all_listings[i], RULES_SEARCH_PAGE['Rooms and Beds'])),
            #price.append(self.extract_element(all_listings[i], RULES_SEARCH_PAGE['Price'])),
            total_price.append(self.extract_element(all_listings[i], RULES_SEARCH_PAGE['Total_Price']))

        data = {'Rank': rank, 'Hotel_Name': name, 'Description':description, 'Rooms_and_beds':rooms_and_beds, 'Total_Price': total_price,'Url':link}
        data['Url'] = ['https://www.airbnb.com' + str(link) for link in data['Url']]
        data= pd.DataFrame(data)
        data['Total_Price'] = data['Total_Price'].astype(str)
        data['Total_Price'] = data['Total_Price'].str.extract('(\d+,?\d*)', expand=False).str.replace(',', '').astype(int)
        return data

    
    
    def apply_conditional_formatting_Eq(self, spreadsheet_name, sheet_name, range_to_format, green_color, red_color):
        
        success = False
        while not success:
            # Open the desired spreadsheet by its name
            try:
                sheet = self.client.open(spreadsheet_name).worksheet(sheet_name)

                # COLOR
                green_rule = ConditionalFormatRule(
                    ranges=[GridRange.from_a1_range(range_to_format, sheet)],
                    booleanRule=BooleanRule(
                        condition=BooleanCondition('NUMBER_GREATER_THAN_EQ', ['0%']),
                        format=CellFormat(textFormat=TextFormat(foregroundColor=Color(*green_color), bold=True))  # Green color
                    ),
                )
                # Define the formatting rule for negative numbers
                red_rule = ConditionalFormatRule(
                    ranges=[GridRange.from_a1_range(range_to_format, sheet)],
                    booleanRule=BooleanRule(
                        condition=BooleanCondition('NUMBER_LESS_THAN_EQ', ['0%']),
                        format=CellFormat(textFormat=TextFormat(foregroundColor=Color(*red_color), bold=True))  # Red color
                    )
                )
                # Apply the formatting rules to the worksheet
                rules = get_conditional_format_rules(sheet)
                # rules.clear()
                rules.append(green_rule)
                rules.append(red_rule)
                rules.save()
                success = True

            except APIError as error:
                
                if error.response.status_code == 429:
                    self.print("Quota exceeded. Retrying in 40 seconds...")
                    time.sleep(40)  # wait for 40 seconds
                else:
                    self.print("APIError occurred: %s" % error)
                    self.print("Retrying in 10 seconds...")
                    time.sleep(10)  # wait for 10 seconds

            except (ConnectionError, Timeout, TooManyRedirects) as error:
                self.print("Connection error occurred: %s" % error)
                self.print("Retrying in 10 seconds...")
                time.sleep(10)  # wait for 10 seconds

            except Exception as error:
                self.print("An error occurred: %s" % error)
                self.print("Retrying in 10 seconds...")
                time.sleep(10)  # wait for 10 seconds


    def apply_conditional_formatting_Num(self, spreadsheet_name, sheet_name, range_to_format, red_color):
        success = False
        while not success:
            
            try:
                sheet = self.client.open(spreadsheet_name).worksheet(sheet_name)
                # Define the formatting rule for negative numbers
                red_rule = ConditionalFormatRule(
                    ranges=[GridRange.from_a1_range(range_to_format, sheet)],
                    booleanRule=BooleanRule(
                        condition=BooleanCondition('NUMBER_GREATER', ['2']),
                        format=CellFormat(textFormat=TextFormat(foregroundColor=Color(*red_color), bold=True))  # Red color
                    )
                )

                # Apply the formatting rules to the worksheet
                rules = get_conditional_format_rules(sheet)
                # rules.clear()
                rules.append(red_rule)
                rules.save()
                success = True
            
            except APIError as error:
                
                if error.response.status_code == 429:
                    self.print("Quota exceeded. Retrying in 40 seconds...")
                    time.sleep(40)  # wait for 40 seconds
                else:
                    self.print("APIError occurred: %s" % error)
                    self.print("Retrying in 10 seconds...")
                    time.sleep(10)  # wait for 10 seconds

            except (ConnectionError, Timeout, TooManyRedirects) as error:
                self.print("Connection error occurred: %s" % error)
                self.print("Retrying in 10 seconds...")
                time.sleep(10)  # wait for 10 seconds

            except Exception as error:
                self.print("An error occurred: %s" % error)
                self.print("Retrying in 10 seconds...")
                time.sleep(10)  # wait for 10 seconds

    def inserting_values(self, spreadsheet_name, sheet_name, row_data):
        success = False
        while not success:
            # Open the desired spreadsheet by its name
            try:
                sheet = self.client.open(spreadsheet_name).worksheet(sheet_name)
                # get the number of rows currently in the sheet
                num_rows = len(sheet.get_all_values())
                # insert the new row of data at the next available row
                sheet.insert_row(row_data, num_rows + 1)
                success = True

            except APIError as error:
                
                if error.response.status_code == 429:
                    self.print("Quota exceeded. Retrying in 40 seconds...")
                    time.sleep(40)  # wait for 40 seconds
                else:
                    self.print("APIError occurred: %s" % error)
                    self.print("Retrying in 10 seconds...")
                    time.sleep(10)  # wait for 10 seconds

            except (ConnectionError, Timeout, TooManyRedirects) as error:
                self.print("Connection error occurred: %s" % error)
                self.print("Retrying in 10 seconds...")
                time.sleep(10)  # wait for 10 seconds

            except Exception as error:
                self.print("An error occurred: %s" % error)
                self.print("Retrying in 10 seconds...")
                time.sleep(10)  # wait for 10 seconds



    def create_new_worksheet(self, spreadsheet_name, worksheet_title):

        spreadsheet_id = self.get_spreadsheet_id(spreadsheet_name)
        sheets_service = build('sheets', 'v4', credentials=self.credentials)

        # Create the new worksheet
        requests = [
            {
                'addSheet': {
                    'properties': {
                        'title': worksheet_title
                    }
                }
            }
        ]
        sheets_service.spreadsheets().batchUpdate(spreadsheetId=spreadsheet_id,
                                                  body={'requests': requests}).execute()

        self.print( f"A new worksheet named '{worksheet_title}' has been added to the Google Spreadsheet with ID '{spreadsheet_id}'")

    def Hotel_10(self,df):
    
        df['Beds'] = df['Rooms_and_beds'].str.extract(r'(\d+) beds')
        df['Bedrooms'] = df['Rooms_and_beds'].str.extract(r'(\d+) bedrooms')
        #df["Total_Price"] = df["Total_Price"].str.extract('(\d+,?\d*)', expand=False).str.replace(',', '').astype(int)
        columns=['Rank', 'Description','Beds','Rooms_and_beds']
        df = df.drop(columns, axis=1)
            
        df['Bedrooms'] = pd.to_numeric(df['Bedrooms'], errors='coerce')
        df['Bedrooms'] = df['Bedrooms'].fillna(0)
        df['Bedrooms'] = df['Bedrooms'].astype(int)
        beds_filter = (df['Bedrooms'].isin([5, 6, 7]))
        df_filtered = df[beds_filter]
        
        # Creating spreadsheet for first 10 Hotels
        df_10 = df_filtered.head(10)

        return df_10

    def inserting_structure_avg_prices(self, spreadsheet_name, sheet_name):
        self.insert_cells_at_start(spreadsheet_name, sheet_name, 8)

        # updating cell size
        ranges_and_sizes = [('B2:D2', 30, 60), ('F2:H2', 30, 60), ('C3:C3', 30, 100), ('G3:G3', 30, 100),
                            ('E3:E3', 30, 30), ('A1:A1', 30, 200)]

        self.update_cell_sizes_avg_price(sheet_name, ranges_and_sizes)

        self.combine_and_update_col(spreadsheet_name, sheet_name, 'B2:D2', "12 Guests")
        self.combine_and_update_col(spreadsheet_name, sheet_name, 'F2:H2', '14 Guests')

        self.update_cell_color(spreadsheet_name,sheet_name, 'B2', 'white')
        self.update_cell_color(spreadsheet_name,sheet_name, 'F2', 'white')

        self.update_spreadsheet_cells_avg_price(sheet_name)

    def calculate_average(self, spreadsheet_name, sheet_name, cell_range):
        spreadsheet_id = self.get_spreadsheet_id(spreadsheet_name)
        # Get data from sheet
        sheet = self.sheets_service.spreadsheets()
        result = sheet.values().get(spreadsheetId=spreadsheet_id, range=sheet_name + '!' + cell_range).execute()
        values = result.get('values', [])

        # Calculate average
        if not values:
            self.print('No data found')
            return None
        num_values = len(values[0])
        total_sum = 0
        for row in values:
            for i in range(num_values):
                if row[i].isdigit():
                    total_sum += int(row[i])
                elif row[i].replace('.', '', 1).isdigit():
                    total_sum += float(row[i])
        avg = total_sum / (len(values) * num_values)
        return avg

    def calculate_average_with_condition(self, spreadsheet_name, sheet_name, condition_range, value_range):
        spreadsheet_id = self.get_spreadsheet_id(spreadsheet_name)
        # Get data from sheet
        sheet = self.sheets_service.spreadsheets()
        condition_result = sheet.values().get(spreadsheetId=spreadsheet_id, range=condition_range).execute()
        value_result = sheet.values().get(spreadsheetId=spreadsheet_id, range=value_range).execute()
        condition_values = condition_result.get('values', [])
        value_values = value_result.get('values', [])

        # Calculate average
        if not condition_values or not value_values:
            self.print('No data found')
            return None
        num_values = len(value_values[0])
        sum = 0
        count = 0
        for i in range(len(condition_values)):
            if condition_values[i][0] in ['5', '6', '7']:
                for j in range(num_values):
                    if value_values[i][j].isdigit():
                        sum += int(value_values[i][j])
                        count += 1
                    elif value_values[i][j].replace('.', '', 1).isdigit():
                        sum += float(value_values[i][j])
                        count += 1
        if count == 0:
            self.print('No matching data found')
            return None
        avg = sum / count
        return avg

    def update_spreadsheet_cells_avg_price(self, spreadsheat_name, sheet_name):
        range_and_values = [
            {
                'range': f'{sheet_name}!B3',
                'values': [['Beds']]
            },
            {
                'range': f'{sheet_name}!C3',
                'values': [['Average Price']]
            },
            {
                'range': f'{sheet_name}!D3',
                'values': [['Link']]
            },
            {
                'range': f'{sheet_name}!A14',
                'values': [['Average cost of stay (5,6,7 BR)']]
            },
            {
                'range': f'{sheet_name}!A15',
                'values': [['Average Price']]
            },
            {
                'range': f'{sheet_name}!A16',
                'values': [['Our Property Price']]
            },
            {
                'range': f'{sheet_name}!F3',
                'values': [['Beds']]
            },
            {
                'range': f'{sheet_name}!G3',
                'values': [['Average Price']]
            },
            {
                'range': f'{sheet_name}!H3',
                'values': [['Link']]
            }
        ]

        body = {
            'value_input_option': 'USER_ENTERED',
            'data': range_and_values,
        }

        response = self.sheets_service.spreadsheets().values().batchUpdate(
            spreadsheetId=self.get_spreadsheet_id(spreadsheat_name),
            body=body,
        ).execute()

        return response

    def update_cell_sizes_avg_price(self, spreadsheat_name, sheet_name, ranges_and_sizes):
    
        sheet_id = self.get_or_create_sheet_id(spreadsheat_name, sheet_name)
        batch_requests = []
        for range_str, row_height, column_width in ranges_and_sizes:
            start_col = column_index_from_string(range_str.split(":")[0][0])
            end_col = column_index_from_string(range_str.split(":")[1][0])
            start_row = int(range_str.split(":")[0][1:])
            end_row = int(range_str.split(":")[1][1:])
            start_col_a1 = get_column_letter(start_col)
            end_col_a1 = get_column_letter(end_col)
            f"{start_col_a1}{start_row}:{end_col_a1}{end_row}"
            request = {
                'updateDimensionProperties': {
                    'range': {
                        'sheetId': sheet_id,
                        'dimension': 'COLUMNS',
                        'startIndex': start_col - 1,
                        'endIndex': end_col,
                    },
                    'properties': {
                        'pixelSize': column_width
                    },
                    'fields': 'pixelSize'
                }
            }
            batch_requests.append(request)
            request = {
                'updateDimensionProperties': {
                    'range': {
                        'sheetId': sheet_id,
                        'dimension': 'ROWS',
                        'startIndex': start_row - 1,
                        'endIndex': end_row,
                    },
                    'properties': {
                        'pixelSize': row_height
                    },
                    'fields': 'pixelSize'
                }
            }
            batch_requests.append(request)

        batch_update_request = {
            'requests': batch_requests
        }
        success = False
        while not success:
            
            try:
                response = self.sheets_service.spreadsheets().batchUpdate(
                    spreadsheetId=self.get_spreadsheet_id(spreadsheat_name),
                    body=batch_update_request).execute()
                success = True   
                self.print(f"Successfully updated cell sizes.")
                success = True
            
            except APIError as error:
                
                if error.response.status_code == 429:
                    self.print("Quota exceeded. Retrying in 40 seconds...")
                    time.sleep(60)  # wait for 40 seconds
                else:
                    self.print("APIError occurred: %s" % error)
                    self.print("Retrying in 10 seconds...")
                    time.sleep(10)  # wait for 10 seconds

            except (ConnectionError, Timeout, TooManyRedirects) as error:
                self.print("Connection error occurred: %s" % error)
                self.print("Retrying in 10 seconds...")
                time.sleep(10)  # wait for 10 seconds

            except Exception as error:
                self.print("An error occurred: %s" % error)
                self.print("Retrying in 10 seconds...")
                time.sleep(10)  # wait for 10 seconds
        
        return response



    def Rank_find(self, df, hotel_name, description, threshold=0.90):
        rank = 0
        if 'Hotel_Name' not in df.columns or 'Description' not in df.columns:
            print ("'Hotel_Name' or 'Description' not found in DataFrame columns")
            rank = 0
            self.print(df)
        else:
            df_hotel = df.loc[(df['Hotel_Name'] == hotel_name) & (df['Description'].apply(lambda x: SequenceMatcher(None, x, description).ratio()) > threshold)]
            if not df_hotel.empty:
                rank = df_hotel['Rank'].values[0]
        return rank

    
    def insert_cells_for_ranking(self, spreadsheet_name, sheet_name):

        self.insert_cells_at_start(spreadsheet_name, sheet_name, 12)
        ranges_and_sizes = [('C1:K1', 50, 70), ('A1:A1', 50, 280), ('B1:B1', 30, 130), ('L1:L1', 50, 40)]
        self.update_cell_sizes_avg_price(spreadsheet_name, sheet_name, ranges_and_sizes)
        today = date.today()
        self.combine_and_update_col(spreadsheet_name, sheet_name, 'B1:K1', str(today))
        self.combine_and_update_col(spreadsheet_name, sheet_name, 'C2:K2', 'Number of Guests')
        self.update_spreadsheet_cells_ranking(spreadsheet_name, sheet_name)

    def insert_dates_into_sheet(self):
        pass

    
    
    def insert_data_of_ranking(self, spreadsheet_name, sheet_name, hotel_name, description,location):
        
        df_filter = pd.read_excel('input_filter.xlsx')
        b=-1
        for a in range(len(df_filter)):
             
            first_row = df_filter.iloc[a]
            
            checkin = str(first_row[0].date())
            checkout = str(first_row[1].date())

            month = first_row[0].date().month
            month = str(calendar.month_name[month])
            filter_date = month + " " + checkin + ' to ' + checkout

            self.get_all_listings(location, checkin, checkout, 8,hotel_name,description)
            
            if self.rank_status:
                b += 1
                date_index = b + 4 + b * 3

                self.update_spreadsheet_cell(spreadsheet_name, sheet_name, 'A' + str(date_index), filter_date)

                self.update_cell_color(spreadsheet_name, sheet_name, 'L1', 'white')
                self.update_cell_color(spreadsheet_name, sheet_name, 'L2', 'white')

                # Conditional Formatting
                self.apply_conditional_formatting_Num(spreadsheet_name, sheet_name,'C' + str(date_index) + ':' + 'K' + str(date_index), (20, 0, 0))
            
                # Updating colortext
                self.update_cell_color(spreadsheet_name, sheet_name, 'L' + str(date_index), 'white')

                self.print(sheet_name)
                self.update_spreadsheet_cell(spreadsheet_name, sheet_name, 'A' + str(date_index + 1),'Avg Price Competitors')
                self.update_spreadsheet_cell(spreadsheet_name, sheet_name, 'A' + str(date_index + 2), 'Our Property Price')
                self.update_spreadsheet_cell(spreadsheet_name, sheet_name, 'A' + str(date_index + 3),'Difference in Price %')
                self.update_cell_color(spreadsheet_name, sheet_name, 'L' + str(date_index + 3), 'white')

                # Conditional Formatting
                self.apply_conditional_formatting_Eq(spreadsheet_name, sheet_name,'C' + str(date_index + 3) + ':' + 'K' + str(date_index + 3), (0, 102, 0), (100, 0, 0))

                # adding chedck box
                self.add_checkbox(spreadsheet_name, sheet_name, 'B' + str(date_index))
                
                
                # Finding today date
                today_date = str(date.today())
                
                
                guest_list = [8,9,10,11,12,13,14,15,16]
                
                ### Need to Move to the seperate thread
                for guest in guest_list:
                    print('Starting Thread for ', guest)
                    sys.stdout.flush()
                    self.updateGuestList(spreadsheet_name, sheet_name, b, guest, location, checkin, checkout, today_date, filter_date, hotel_name, description)

            else:
                self.print('hotel not found skipping this filter.')
                continue

    ##MAIN FUNCTION TO GET DATA AND FLUSH IT
    @background
    def updateGuestList(self, spreadsheet_name, sheet_name, b, guest, location, checkin, checkout, today_date, filter_date, hotel_name, description):
        df = self.get_all_listings(location, checkin, checkout, guest,hotel_name,description)
        rank = self.Rank_find(df, hotel_name,description)
        if rank > 0:
            price = self.Property_Price(df, rank, hotel_name,description)
        else:
            price = 0
        
        if len(df) < 36:
            # Finding Average price for the hotesl of 5,6,7 beds
            listings=self.get_all_listings_Hotel_10(location, checkin, checkout, guest)
            total_10 = self.dataframe_buliding(listings,self.RULES_SEARCH_PAGE)
            df_10 = self.Hotel_10(total_10)
            avg = int(df_10['Total_Price'].mean())
        else:
            df_10 = self.Hotel_10(df)
            avg = int(df_10['Total_Price'].mean())
        
        if guest == 8:
            
            
            self.update_spreadsheet_cell(spreadsheet_name, sheet_name, 'C' + str(b + 4 + b * 3), str(rank))
            self.update_spreadsheet_cell(spreadsheet_name, sheet_name, 'C' + str(b + 5 + b * 3), str('$'+str(avg)))
            self.update_spreadsheet_cell(spreadsheet_name, sheet_name, 'C' + str(b + 6 + b * 3), str('$'+str(price)))
            self.update_spreadsheet_cell(spreadsheet_name, sheet_name, 'C' + str(b + 7 + b * 3),
                                        str(int(((price - avg) / avg) * 100)) + '%')

            if rank > 2:
                data = [today_date,filter_date, hotel_name, str(guest), str('$'+str(price)),str('$'+str(avg)),str(rank)]
                self.inserting_values(spreadsheet_name, 'Rank Below 2 Summary', data)



        elif guest == 9:
            self.update_spreadsheet_cell(spreadsheet_name, sheet_name, 'D' + str(b + 4 + b * 3), str(rank))
            self.update_spreadsheet_cell(spreadsheet_name, sheet_name, 'D' + str(b + 5 + b * 3), str('$'+str(avg)))
            self.update_spreadsheet_cell(spreadsheet_name, sheet_name, 'D' + str(b + 6 + b * 3), str('$'+str(price)))
            self.update_spreadsheet_cell(spreadsheet_name, sheet_name, 'D' + str(b + 7 + b * 3),
                                        str(int(((price - avg) / avg) * 100)) + '%')
            if rank > 2:
                data = [today_date,filter_date, hotel_name, str(guest), str('$'+str(price)),str('$'+str(avg)),str(rank)]
                self.inserting_values(spreadsheet_name, 'Rank Below 2 Summary', data)

        elif guest == 10:
            self.update_spreadsheet_cell(spreadsheet_name, sheet_name, 'E' + str(b + 4 + b * 3), str(rank))
            self.update_spreadsheet_cell(spreadsheet_name, sheet_name, 'E' + str(b + 5 + b * 3), str('$'+str(avg)))
            self.update_spreadsheet_cell(spreadsheet_name, sheet_name, 'E' + str(b + 6 + b * 3), str('$'+str(price)))
            self.update_spreadsheet_cell(spreadsheet_name, sheet_name, 'E' + str(b + 7 + b * 3),
                                        str(int(((price - avg) / avg) * 100)) + '%')

            if rank > 2:
                data = [today_date,filter_date, hotel_name, str(guest), str('$'+str(price)),str('$'+str(avg)),str(rank)]
                self.inserting_values(spreadsheet_name, 'Rank Below 2 Summary', data)

        elif guest == 11:
            self.update_spreadsheet_cell(spreadsheet_name, sheet_name, 'F' + str(b + 4 + b * 3), str(rank))
            self.update_spreadsheet_cell(spreadsheet_name, sheet_name, 'F' + str(b + 5 + b * 3), str('$'+str(avg)))
            self.update_spreadsheet_cell(spreadsheet_name, sheet_name, 'F' + str(b + 6 + b * 3), str('$'+str(price)))
            self.update_spreadsheet_cell(spreadsheet_name, sheet_name, 'F' + str(b + 7 + b * 3),
                                        str(int(((price - avg) / avg) * 100)) + '%')

            if rank > 2:
                data = [today_date,filter_date, hotel_name, str(guest), str('$'+str(price)),str('$'+str(avg)),str(rank)]
                self.inserting_values(spreadsheet_name, 'Rank Below 2 Summary', data)

        elif guest == 12:

            self.update_spreadsheet_cell(spreadsheet_name, sheet_name, 'G' + str(b + 4 + b * 3), str(rank))
            self.update_spreadsheet_cell(spreadsheet_name, sheet_name, 'G' + str(b + 5 + b * 3), str('$'+str(avg)))
            self.update_spreadsheet_cell(spreadsheet_name, sheet_name, 'G' + str(b + 6 + b * 3), str('$'+str(price)))
            self.update_spreadsheet_cell(spreadsheet_name, sheet_name, 'G' + str(b + 7 + b * 3),
                                        str(int(((price - avg) / avg) * 100)) + '%')

            if rank > 2:
                data = [today_date,filter_date, hotel_name, str(guest), str('$'+str(price)),str('$'+str(avg)),str(rank)]
                self.inserting_values(spreadsheet_name, 'Rank Below 2 Summary', data)


        elif guest == 13:
            self.update_spreadsheet_cell(spreadsheet_name, sheet_name, 'H' + str(b + 4 + b * 3), str(rank))
            self.update_spreadsheet_cell(spreadsheet_name, sheet_name, 'H' + str(b + 5 + b * 3), str('$'+str(avg)))
            self.update_spreadsheet_cell(spreadsheet_name, sheet_name, 'H' + str(b + 6 + b * 3), str('$'+str(price)))
            self.update_spreadsheet_cell(spreadsheet_name, sheet_name, 'H' + str(b + 7 + b * 3),
                                        str(int(((price - avg) / avg) * 100)) + '%')

            if rank > 2:
                data = [today_date,filter_date, hotel_name, str(guest), str('$'+str(price)),str('$'+str(avg)),str(rank)]
                self.inserting_values(spreadsheet_name, 'Rank Below 2 Summary', data)

        elif guest == 14:
            self.update_spreadsheet_cell(spreadsheet_name, sheet_name, 'I' + str(b + 4 + b * 3), str(rank))
            self.update_spreadsheet_cell(spreadsheet_name, sheet_name, 'I' + str(b + 5 + b * 3), str('$'+str(avg)))
            self.update_spreadsheet_cell(spreadsheet_name, sheet_name, 'I' + str(b+ 6 + b * 3), str('$'+str(price)))
            self.update_spreadsheet_cell(spreadsheet_name, sheet_name, 'I' + str(b + 7 + b * 3), str(int(((price - avg) / avg) * 100)) + '%')

            if rank > 2:
                data = [today_date,filter_date, hotel_name, str(guest), str('$'+str(price)),str('$'+str(avg)),str(rank)]
                self.inserting_values(spreadsheet_name, 'Rank Below 2 Summary', data)

        elif guest == 15:
            
            self.update_spreadsheet_cell(spreadsheet_name, sheet_name, 'J' + str(b + 4 + b * 3), str(rank))
            self.update_spreadsheet_cell(spreadsheet_name, sheet_name, 'J' + str(b + 5 + b * 3), str('$'+str(avg)))
            self.update_spreadsheet_cell(spreadsheet_name, sheet_name, 'J' + str(b + 6 + b * 3), str('$'+str(price)))
            self.update_spreadsheet_cell(spreadsheet_name, sheet_name, 'J' + str(b + 7 + b * 3), str(int(((price - avg) / avg) * 100)) + '%')

            if rank > 2:
                data = [today_date,filter_date, hotel_name, str(guest), str('$'+str(price)),str('$'+str(avg)),str(rank)]
                self.inserting_values(spreadsheet_name, 'Rank Below 2 Summary', data)

        elif guest == 16:
            self.update_spreadsheet_cell(spreadsheet_name, sheet_name, 'K' + str(b + 4 + b * 3), str(rank))
            self.update_spreadsheet_cell(spreadsheet_name, sheet_name, 'K' + str(b + 5 + b * 3), str('$'+str(avg)))
            self.update_spreadsheet_cell(spreadsheet_name, sheet_name, 'K' + str(b + 6 + b * 3), str('$'+str(price)))
            self.update_spreadsheet_cell(spreadsheet_name, sheet_name, 'K' + str(b + 7 + b * 3), str(int(((price - avg) / avg) * 100)) + '%')
            
            if rank > 2:
                data = [today_date,filter_date, hotel_name, str(guest), str('$'+str(price)),str('$'+str(avg)),str(rank)]
                self.inserting_values(spreadsheet_name, 'Rank Below 2 Summary', data)



    def insert_cells_for_price_log(self, spreadsheet_name, sheet_name):

        self.insert_cells_at_start(spreadsheet_name, sheet_name, 8)
        ranges_and_sizes = [('A2:A2', 60, 250), ('B2:G2', 60, 150), ('H2:H2', 60, 30)]
        self.update_cell_sizes_avg_price(spreadsheet_name, sheet_name, ranges_and_sizes)
        self.update_cell_color(spreadsheet_name, sheet_name, 'H1','white')
        self.update_cell_color(spreadsheet_name, sheet_name, 'H2','white')
    
    
    def inserting_columns_names_priceLog(self, spreadsheet_name, sheet_name,hotel_name):
        retry_delay = 10
        #self.sheets_service = build('sheets', 'v4', credentials=self.credentials)
        range_and_values = [
            {
                'range': f'{sheet_name}!A2',
                'values': [['Date Change Made\n'+ str('( '+(hotel_name)+' )')]]
            },
            {
                'range': f'{sheet_name}!B2',
                'values': [['Price Change in Airbnb']]
            },
            {
                'range': f'{sheet_name}!C2',
                'values': [['% price change in Airbnb']]
            },
            {
                'range': f'{sheet_name}!D2',
                'values': [['Rank Before (for 12 guest)']]
            },
            {
                'range': f'{sheet_name}!E2',
                'values': [['Rank After (for 12 guest)']]
            },
            {
                'range': f'{sheet_name}!F2',
                'values': [['Rank Change After Price Change?']]
            },
            {
                'range': f'{sheet_name}!G2',
                'values': [['Action for Next Day']]
            }
        ]

        body = {
            'value_input_option': 'USER_ENTERED',
            'data': range_and_values,
        }
        success = False
        while not success:
            try:
                response = self.sheets_service.spreadsheets().values().batchUpdate(
                    spreadsheetId=self.get_spreadsheet_id(spreadsheet_name),
                    body=body,
                ).execute()
                success = True

            except APIError as error:
                
                if error.response.status_code == 429:
                    self.print("Quota exceeded. Retrying in 60 seconds...")
                    time.sleep(retry_delay)
                    retry_delay *= 2   # wait for 40 seconds
                else:
                    self.print("APIError occurred: %s" % error)
                    self.print("Retrying in 10 seconds...")
                    time.sleep(10)  # wait for 10 seconds

            except Exception as error:
                self.print("An error occurred: %s" % error)
                self.print("Retrying in 10 seconds...")
                time.sleep(10)  # wait for 10 seconds

        return response
    
    def search_value_in_column(self,spreadsheet_name, worksheet_name, column, value):
        
        retry_delay = 10
        success = False
        while not success:
            try:
                sheet = self.client.open(spreadsheet_name).worksheet(worksheet_name)

                # Get all values in the specified column
                column_values = sheet.col_values(column)
                # Search for the value in the column
                if value in column_values:
                    row = column_values.index(value) + 1
                    success = True
                    return row
                else:
                    success = True
                    return None
                
            except APIError as error:
                
                if error.response.status_code == 429:
                    self.print("Quota exceeded. Retrying in 60 seconds...")
                    time.sleep(retry_delay)
                    retry_delay *= 2   # wait for 40 seconds
                else:
                    self.print("APIError occurred: %s" % error)
                    self.print("Retrying in 10 seconds...")
                    time.sleep(10)  # wait for 10 seconds
            
            except Exception as error:
                self.print("An error occurred in 'search_value_in_column ': %s" % error)
                self.print("Retrying in 10 seconds...")
                time.sleep(10)  # wait for 10 seconds
        return row
            
    def insert_dates_into_column(self, spreadsheet_name, worksheet_name, column):
        retry_delay = 10
        self.print('Inserting dates in column A1')
        df_filter = pd.read_excel('input_filter.xlsx')
        dates= []
        
        for a in range(len(df_filter)):
            
            first_row = df_filter.iloc[a]
            checkin = str(first_row[0].date())
            checkout = str(first_row[1].date())

            month = first_row[0].date().month
            month = str(calendar.month_name[month])
            filter_date = month + " " + checkin + ' to ' + checkout
            dates.append(filter_date)

        success = False
        while not success:
            try:
                sheet = self.client.open(spreadsheet_name).worksheet(worksheet_name)
                # Determine the range of cells where to insert the values
                start_cell = f'{column}2'
                end_cell = f'{column}{len(dates)+1}'
                range_to_update = f'{start_cell}:{end_cell}'

                # Insert the values into the specified range
                sheet.update(range_to_update, [[value] for value in dates])
                success = True
            
            except APIError as error:
                
                if error.response.status_code == 429:
                    self.print("Quota exceeded. Retrying in 60 seconds...")
                    time.sleep(retry_delay)
                    retry_delay *= 2   # wait for 40 seconds  # wait for 40 seconds
                else:
                    self.print("APIError occurred: %s" % error)
                    self.print("Retrying in 10 seconds...")
                    time.sleep(10)  # wait for 10 seconds

            except Exception as error:
                self.print("An error occurred: %s" % error)
                self.print("Retrying in 10 seconds...")
                time.sleep(10)  # wait for 10 seconds

        self.print("Dates Inserted")
    
    def insert_values_in_range(self, spreadsheet_name, worksheet_name, cell_range, values):
        retry_delay = 10
        success = False
        while not success:
            try:
                # Prepare the update request
                start_cell, end_cell = cell_range.split(':')
                start_row = int(start_cell[1:])
                end_row = int(end_cell[1:])
                num_rows = end_row - start_row + 1
                num_columns = ord(end_cell[0]) - ord(start_cell[0]) + 1

                value_range = {
                    'range': f"{worksheet_name}!{cell_range}",
                    'majorDimension': 'ROWS',
                    'values': [values[row * num_columns: (row + 1) * num_columns] for row in range(num_rows)],
                }

                # Send the update request
                request = self.sheets_service.spreadsheets().values().update(
                    spreadsheetId=self.get_spreadsheet_id(spreadsheet_name),
                    range=f"{worksheet_name}!{cell_range}",
                    valueInputOption='USER_ENTERED',
                    body=value_range
                )
                response = request.execute()
                success = True

            except (HttpError, APIError) as error:
                if error.resp.status == 429:
                    self.print("Quota exceeded. Retrying in 60 seconds...")
                    time.sleep(retry_delay)
                    retry_delay *= 2
                else:
                    self.print("HttpError occurred: %s" % error)
                    self.print("Retrying in 10 seconds...")
                    time.sleep(10)

            except Exception as error:
                self.print("An error occurred: %s" % error)
                self.print("Retrying in 10 seconds...")
                time.sleep(10)

    
    def get_cell_value(self, spreadsheet_name, worksheet_name, cell_reference):
        retry_delay = 10
        success = False
        while not success:
            try:
                # Select the worksheet
                sheet = self.client.open(spreadsheet_name).worksheet(worksheet_name)

                # Check if the cell reference is a single cell or a range
                if ':' in cell_reference:
                    # Range of cells
                    cell_range = sheet.range(cell_reference)
                    num_columns = int(cell_reference.split(':')[1][1:]) - int(cell_reference.split(':')[0][1:]) + 1
                    values = [cell.value for cell in cell_range]
                    if all(element == '' for element in values):
                        self.print('Rank values not found! trying again! ')
                        continue
                else:
                    # Single cell
                    cell_value = sheet.acell(cell_reference).value
                    values = cell_value 
                success = True
            
            except APIError as error:
                if error.response.status_code == 429:
                    self.print("Quota exceeded. Retrying in 60 seconds...")
                    time.sleep(retry_delay)
                    retry_delay *= 2   # wait for 40 seconds
                else:
                    self.print("APIError occurred: %s" % error)
                    self.print("Retrying in 10 seconds...")
                    time.sleep(10)  # wait for 10 seconds
            except Exception as error:
                self.print("An error occurred: %s" % error)
                self.print("Retrying in 10 seconds...")
                time.sleep(10)  # wait for 10 seconds
            
        return values


    
    def convert_price_to_int(self, string_value):
        if string_value is None:
            return 0  # or handle it in another way
        numeric_string = ''.join(filter(str.isdigit, string_value))
        return int(numeric_string)
    
    def price_change_log(self,spreadsheet_name,hotel_ranking_sheet,worksheet_name,hotel_name):
        
        self.insert_cells_for_price_log(spreadsheet_name,worksheet_name)
        today = date.today()
        self.combine_and_update_col(spreadsheet_name, worksheet_name, 'A1:G1', str(today))
        self.inserting_columns_names_priceLog(spreadsheet_name,worksheet_name,hotel_name)
        #self.insert_dates_into_column(spreadsheet_name, worksheet_name, 'A')
        
        df_filter = pd.read_excel('input_filter.xlsx')
        data = []
        b = -1
        for a in range(len(df_filter)):
            
            date_search = self.get_cell_value(spreadsheet_name,hotel_ranking_sheet,'A'+str(a + 4 + a * 3))
            previous_row = self.search_value_in_column(spreadsheet_name,hotel_ranking_sheet,13, date_search)
        
            if previous_row != None:
                old_price = self.get_cell_value(spreadsheet_name,hotel_ranking_sheet,'S'+str(previous_row+2))
                b += 1
            else:
                self.print('date not found continue..')
                continue
            
            date_row =  a + 4 + a * 3

            old_price = self.convert_price_to_int(old_price)
            self.print('Previous Price:', old_price)
            
            today_price = self.get_cell_value(spreadsheet_name,hotel_ranking_sheet,'G'+str(date_row+2))
            today_price = self.convert_price_to_int(today_price)
            self.print('Today Price:', today_price)
    
            rank_before = self.get_cell_value(spreadsheet_name,hotel_ranking_sheet,'S'+ str(previous_row))
            self.print('rank_before: ', rank_before)
            
            rank_after= self.get_cell_value(spreadsheet_name,hotel_ranking_sheet,'G'+str(date_row))
            self.print('rank_after: ', rank_after)

            if rank_before != rank_after:
                rank_change= 'Yes'
            else:
                rank_change= 'No'
            self.print(rank_change)    
            
            percentage_change = round((((today_price - old_price) / old_price) * 100),2)
            self.print(percentage_change)

            self.print('F'+str(date_row)+':'+'K'+str(date_row))
            current_ranks_11_to_16 = self.get_cell_value(spreadsheet_name, hotel_ranking_sheet, 'F'+str(date_row)+':'+'K'+str(date_row))
            self.print('R'+str(previous_row)+':'+'W'+str(previous_row))
            previous_ranks_11_to_16 = self.get_cell_value(spreadsheet_name, hotel_ranking_sheet, 'R'+str(previous_row)+':'+'W'+str(previous_row))
            self.print('current ranks',current_ranks_11_to_16)
            self.print('Previous ranks',previous_ranks_11_to_16)
            
            current_all_ones = all(int(element) == 1 for element in current_ranks_11_to_16)
            previous_all_ones = all(int(element) == 1 for element in previous_ranks_11_to_16)
            if current_all_ones:
                if  previous_all_ones:
                    next_change = 1
                else:
                    next_change= "Don't Change!"
            else:
                next_change = -1
            
            data = [date_search, today_price,percentage_change,rank_before, rank_after,rank_change,next_change]
            self.print(data)
            self.insert_values_in_range(spreadsheet_name, worksheet_name, 'A'+str(b+3)+':'+'G'+str(b+3)  ,data)
            

        
        #inserting dates
        
        #self.insert_values_into_column(spreadsheet_name, worksheet_name, 'A', dates)
        #getting Values from sheet
        
        #cell_value = self.get_cell_value(spreadsheet_name,worksheet_name,'G4')
        #self.print(type(cell_value))
        #self.print(cell_value)
        
    def update_spreadsheet_range(self, spreadsheet_name, sheet_name, update_range, values):
        body = {'values': values}
        result = self.sheets_service.spreadsheets().values().update(
            spreadsheetId=self.get_spreadsheet_id(spreadsheet_name), range=sheet_name + '!' + update_range,
            valueInputOption='RAW', body=body).execute()

    def insert_data_avg_price(self, spreadsheet_name, sheet_name, df_filters,description):
        for a in range(len(df_filters)):

            first_row = df_filters.iloc[a]
            location = first_row[0]
            checkin = str(first_row[1].date())
            checkout = str(first_row[2].date())

            # insering cells the spreadsheet
            self.inserting_structure_avg_prices(spreadsheet_name,sheet_name)

            month = first_row[2].date().month
            month = str(calendar.month_name[month])
            date1 = month + " " + checkin + ' to ' + checkout

            # Showing filter on the first column
            self.combine_and_update_col(spreadsheet_name, sheet_name, 'B1:H1', str(date1))

            guest_list = [12, 14]

            for guest in guest_list:

                listings = self.get_all_listings(location, checkin, checkout, guest)
                df = self.dataframe_buliding(listings, self.RULES_SEARCH_PAGE)

                # Creating spreadsheet for first 10 Hotels
                df_10 = self.Hotel_10(df)

                beds = df_10['Bedrooms'].values
                total_price = df_10['Total_Price'].values
                Url = df_10['Url'].values

                if guest == 12:
                    # time.sleep(2)cls
                    for i in range(len(df_10)):
                        update_range = 'B4:B' + str(len(df_10) + 3)
                        values = [[str(bed)] for bed in beds]
                        self.update_spreadsheet_range(spreadsheet_name,sheet_name, update_range, values)

                        # update_spreadsheet_cell(sheet_name,'B'+str(i+4), str(beds[i]))

                    for i in range(len(df_10)):
                        update_range = 'C4:C' + str(len(df_10) + 3)
                        values = [[str(Total)] for Total in total_price]
                        self.update_spreadsheet_range(spreadsheet_name,sheet_name, update_range, values)

                        # update_spreadsheet_cell(sheet_name,'C'+str(i+4), str(Total_Price[i]))

                    for i in range(len(df_10)):
                        update_range = 'D4:D' + str(len(df_10) + 3)
                        values = [[str(url)] for url in Url]
                        self.update_spreadsheet_range(spreadsheet_name,sheet_name, update_range, values)

                        # update_spreadsheet_cell(sheet_name,'D'+str(i+4), str(Url[i]))

                    # Finding average prices
                    total_avg = self.calculate_average(spreadsheet_name, sheet_name, 'C4:C13')
                    avg = self.calculate_average_with_condition(spreadsheet_name, sheet_name, 'B4:B13', 'C4:C13')
                    self.update_spreadsheet_cell(spreadsheet_name,sheet_name, 'C14', str(avg))
                    self.update_spreadsheet_cell(spreadsheet_name,sheet_name, 'C15', str(total_avg))
                    price = self.Property_Price(df, 'Villa in Hollywood',description)
                    self.update_spreadsheet_cell(spreadsheet_name,sheet_name, 'C16', str(price))



                elif guest == 14:
                    # time.sleep(2)
                    for i in range(len(df_10)):
                        update_range = 'F4:F' + str(len(df_10) + 3)
                        values = [[str(bed)] for bed in beds]
                        self.update_spreadsheet_range(spreadsheet_name,sheet_name, update_range, values)

                        # update_spreadsheet_cell(sheet_name,'F'+str(i+4), str(beds[i]))
                    for i in range(len(df_10)):
                        update_range = 'G4:G' + str(len(df_10) + 3)
                        values = [[str(Total)] for Total in total_price]
                        self.update_spreadsheet_range(spreadsheet_name,sheet_name, update_range, values)

                        # update_spreadsheet_cell(sheet_name,'G'+str(i+4), str(Total_Price[i]))
                    for i in range(len(df_10)):
                        update_range = 'H4:H' + str(len(df_10) + 3)
                        values = [[str(url)] for url in Url]
                        self.update_spreadsheet_range(spreadsheet_name,sheet_name, update_range, values)

                        # update_spreadsheet_cell(sheet_name,'H'+str(i+4), str(Url[i]))

                    # Finding average prices
                    total_avg = self.calculate_average(spreadsheet_name, sheet_name, 'G4:G14')
                    avg = self.calculate_average_with_condition(spreadsheet_name, sheet_name, 'F4:F13', 'G4:G13')
                    self.update_spreadsheet_cell(spreadsheet_name,sheet_name, 'G14', str(avg))
                    self.update_spreadsheet_cell(spreadsheet_name,sheet_name, 'G15', str(total_avg))
                    price = self.Property_Price(df, 'Villa in Hollywood',description)
                    self.update_spreadsheet_cell(spreadsheet_name,sheet_name, 'G16', str(price))

    def main(self,spreadsheet_name,sheet_name,description,location,price_log_sheet):
        # Ranking
        
        self.insert_cells_for_ranking(spreadsheet_name, sheet_name)
        self.insert_data_of_ranking(spreadsheet_name, sheet_name,sheet_name,description,location)
        
        self.print("Inserting Values into Price log sheet..")
        self.price_change_log(spreadsheet_name,sheet_name,price_log_sheet,sheet_name)

        # Avgerage Prices
        # df_filters= pd.read_excel('input_filter.xlsx')
        # df_filters.drop(columns=["Guests"], inplace=True)

        # Avgerage Prices
        # self.insert_data_avg_price('Ranking Data','Average Prices of hotels', df_filters)


# if __name__ == "__main__":
#     # Ranking
#     insertion = DataInsertion()
#     insertion.main()
